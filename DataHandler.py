import pandas as pd
import openpyxl as pyxl
import numpy as np

import dbconnect as db

years_included = ['16']
newStats = 'NewStats'


def getdataframefromfile(fileTitle):
    filepath = 'data/spreadsheets/' + fileTitle + '.xlsx'

    rangevalues = np.array([])

    preCreateStatsSheet(newStats, filepath)
    activeSheets = findActiveSheets(filepath)

    persons = np.empty([0, 4])
    for sheetName in activeSheets:
        currentsheet = getColumnNamesRange(sheetName, filepath)
        persons = np.concatenate((persons, currentsheet), axis=0)

    analyzePersons(persons, filepath)

    data = pd.read_excel(filepath)
    return data


def as_text(value):
    if value is None:
        return ""
    return str(value)


def preCreateStatsSheet(sheetname, filepath): #realized
    wb = pyxl.load_workbook(filepath)

    statsSheet = wb.create_sheet(sheetname, 0)
    statsSheet.title = sheetname
    statsSheet.append(["Имя", "Победы", "Ничьи", "Поражения", "Всего Игр", "Коэффициент побед", "Коэффициент очков"])

    wb.save(filename=filepath)


def findActiveSheets(filepath):
    wb = pyxl.load_workbook(filepath, data_only=True)

    ws = wb.get_sheet_names()

    activeSheets = list()
    for sheet in ws:
        if (isYearOK(sheet)) and sheet.find('Stats') == -1 and sheet.find('год') == -1:
            activeSheets.append(sheet)
    print("Total number of all sheets = " + str(activeSheets.__len__()))

    print("This is pages:")
    for sheetName in activeSheets:
        print (sheetName)

    return activeSheets


def isYearOK(year):
    if (years_included == None):
        raise 'No selected years Error'

    for years in years_included:
        if (year[-2::1].rfind(years) != -1):
            return True

    return False


def getColumnNamesRange(currentSheet, filepath):
    wb = pyxl.load_workbook(filepath, data_only=True)
    ws = wb.get_sheet_by_name(currentSheet)

    range = []
    rangestats = np.empty([0, 4])
    startrow = 4000 #example max people
    scorerow = 0
    scorecell = ws['A1']
    stringscore = ''
    scorearray = []
    rows = ws.max_row
    maxrows = None
    titledata = ws.title
    titleyear = (int)(titledata[-2:])
    titlemonth = (int)(titledata.replace(' ', '')[:-2])
    lastcolumn = 0
    index = 0
    valuelist = []
    wingames = 0
    drawgames = 0
    totalgames = 0
    cellcolor = ''
    bluecolor = 'FF0000FF'
    redcolor =  'FFFF0000'
    whitecolor = '00000000' # or 'FFFFFFFF'
    greencolor =  'FF6AA84F'
    yellowcolor = 'FFFFFF00'
    darkgreencolor = 'FF38761D'
    # case 1 Current date - 12/13 (First column)
    if((titleyear >= 13 and titlemonth > 12) or titleyear >= 14):
        for cell in ws[1]:
            if (cell.value == 'Оплата'):
                lastcolumn = cell.column
                break
        for cell in ws['A']:
            if(cell.value == 'Аренда'):
                startrow = (int)(cell.row)
            if(cell.value == 'счет' or cell.value == 'счёт'):
                scorecell = cell
                scorerow = cell.row
            if(cell.value == '' or cell.value == None):
                break
            if((int)(cell.row) > startrow and cell.value != 'Guests ' and cell.value != 'Guests'):

                for cellscore in ws[scorecell.row]:
                    if(cellscore.fill.start_color.index == yellowcolor):
                        break
                    currentcell = ws[(str)(cellscore.column) + (str)(cell.row)]
                    if (cellscore.column > 'B' and cellscore.column < lastcolumn and cellscore.value != None and currentcell.value != None and currentcell.value != ''):
                        cellcolor = cellscore.fill.start_color.index
                        currentcellcolor = currentcell.fill.start_color.index
                        if(cellcolor == whitecolor):
                            totalgames += 1
                            drawgames += 1
                        if (cellcolor == redcolor):
                            if (currentcellcolor == redcolor):
                                wingames += 1
                                totalgames += 1
                            elif(currentcellcolor == bluecolor):
                                totalgames += 1
                        elif (cellcolor == bluecolor):
                            if (currentcellcolor == bluecolor):
                                wingames += 1
                                totalgames += 1
                            elif(currentcellcolor == redcolor):
                                totalgames += 1

                range.append(cell.value)
                valuelist.append(cell.value)
                valuelist.append(wingames)
                valuelist.append(drawgames)
                valuelist.append(totalgames)
                rangestats = np.vstack((rangestats, valuelist))

                wingames = 0
                totalgames = 0
                drawgames = 0
                valuelist = []

    #Case 2 12/13 - 04/13 Second column(First column have numeration)
    elif(titleyear == 13 and (titlemonth >= 4 and titlemonth <= 12)):
        for cell in ws['B']:
            if (cell.value == 'счет' or cell.value == 'счёт'):
                scorecell = cell
                scorerow = cell.row
            if (cell.value == '' or cell.value == None):
                break
            if(cell.value == 'Аренда'):
                startrow = (int)(cell.row)
            if(cell.value == '' or cell.value == None):
                break
            if((int)(cell.row) > startrow and cell.value != 'Guests ' and cell.value != 'Guests'):

                for cellscore in ws[scorecell.row]:
                    if(cellscore.fill.start_color.index == yellowcolor):
                        break
                    currentcell = ws[(str)(cellscore.column) + (str)(cell.row)]
                    if (cellscore.column > 'B' and cellscore.value != None and currentcell.value != None and currentcell.value != ''):
                        cellcolor = cellscore.fill.start_color.index
                        currentcellcolor = currentcell.fill.start_color.index
                        if(cellcolor == darkgreencolor):
                            totalgames += 1
                            drawgames += 1
                        if (cellcolor == redcolor):
                            if (currentcellcolor == redcolor):
                                wingames += 1
                                totalgames += 1
                            elif(currentcellcolor == bluecolor):
                                totalgames += 1
                        elif (cellcolor == bluecolor):
                            if (currentcellcolor == bluecolor):
                                wingames += 1
                                totalgames += 1
                            elif(currentcellcolor == redcolor):
                                totalgames += 1
                range.append(cell.value)
                valuelist.append(cell.value)
                valuelist.append(wingames)
                valuelist.append(drawgames)
                valuelist.append(totalgames)
                rangestats = np.vstack((rangestats, valuelist))

                wingames = 0
                totalgames = 0
                drawgames = 0
                valuelist = []

    else:
        scorecell = ws['A1']
        for cell in ws['B']:
            if(cell.value == 'счет'):
                scorecell = cell
                startrow = 1
                break
        for cell in ws['B']:
            if(cell.value == '' or cell.value == None or cell.value == 'Сумма' or cell.value == 'счет'):
                break
            if((int)(cell.row) >= startrow and cell.value != 'Guests ' and cell.value != 'Guests'):

                for cellscore in ws[scorecell.row]:
                    if(cellscore.fill.start_color.index == yellowcolor):
                        break
                    currentcell = ws[(str)(cellscore.column) + (str)(cell.row)]
                    if (cellscore.column > 'B' and currentcell.value != None and currentcell.value != ''):
                        cellcolor = cellscore.fill.start_color.index
                        currentcellcolor = currentcell.fill.start_color.index
                        if(cellcolor == whitecolor):
                            totalgames += 1
                            drawgames += 1
                        if (cellcolor == redcolor):
                            if (currentcellcolor == redcolor):
                                wingames += 1
                                totalgames += 1
                            elif(currentcellcolor == bluecolor):
                                totalgames += 1
                        elif (cellcolor == bluecolor):
                            if (currentcellcolor == bluecolor):
                                wingames += 1
                                totalgames += 1
                            elif(currentcellcolor == redcolor):
                                totalgames += 1
                range.append(cell.value)
                valuelist.append(cell.value)
                valuelist.append(wingames)
                valuelist.append(drawgames)
                valuelist.append(totalgames)
                rangestats = np.vstack((rangestats, valuelist))

                wingames = 0
                drawgames = 0
                totalgames = 0
                valuelist = []
    return rangestats


def analyzePersons(persons, filePath):
    # Open file
    wb = pyxl.load_workbook(filePath)
    statsSheet = wb.get_sheet_by_name(newStats)

    # Process the list of persons
    i = 0
    k = 0

    # for person in persons:
    while (k < persons.shape[0]):
        # One sheet info
        footballplayername = persons[k,0]

        # Sum player's stat and delete useless info
        deletingindexes = list()
        for j in range(i + 1, persons.shape[0]):
            if (footballplayername == persons[j, 0]):
                persons[i, 1] = int(persons[i, 1]) + int(persons[j, 1])
                persons[i, 2] = int(persons[i, 2]) + int(persons[j, 2])
                persons[i, 3] = int(persons[i, 3]) + int(persons[j, 3])
                deletingindexes.append(j)

        persons = np.delete(persons, deletingindexes, 0)

        footballplayerwins = int(persons[k,1])
        footballplayerdraws = int(persons[k,2])
        footballplayernumofgames = int(persons[k,3])

        try:
            footballplayervictoryrate = str(int((footballplayerwins / footballplayernumofgames) * 100)) + '%'
            footballplayerscorerate = str(
                int(((footballplayerwins * 3 + footballplayerdraws) / (footballplayernumofgames * 3)) * 100)) + '%'

        except ZeroDivisionError:
            footballplayervictoryrate = '0'
            footballplayerscorerate = '0'

        # Set info to the stats
        row = i + 2
        statsSheet.cell(row=row, column=1).value = footballplayername
        statsSheet.cell(row=row, column=2).value = footballplayerwins
        statsSheet.cell(row=row, column=3).value = footballplayerdraws
        statsSheet.cell(row=row, column=4).value = footballplayernumofgames - footballplayerwins - footballplayerdraws
        statsSheet.cell(row=row, column=5).value = footballplayernumofgames
        statsSheet.cell(row=row, column=6).value = footballplayervictoryrate
        statsSheet.cell(row=row, column=7).value = footballplayerscorerate

        i += 1
        k += 1

    wb.save(filename=filePath)


def testdb():
    c, conn = db.connection()
    c.execute('SELECT * FROM Players ORDER BY PlayerID')
    data = c.fetchall()
    dataframe = pd.DataFrame(list(data), columns=['ID', 'Имя игрока'])
    return dataframe